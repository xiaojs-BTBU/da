import pandas as pd
import io
import sys

# 模拟一个 .xls 文件（旧版 Excel）的内容
# 由于没有实际文件，我们测试 openpyxl 是否可以读取 .xls 扩展名的文件
# 使用一个虚拟的 BytesIO 对象
from openpyxl import load_workbook

print("测试 openpyxl 读取 .xls 文件的能力")
try:
    # 创建一个新的工作簿
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'test'
    # 保存到字节流
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    # 尝试用 openpyxl 加载
    wb2 = load_workbook(buffer)
    print("openpyxl 可以读取 .xlsx 格式的内存工作簿")
except Exception as e:
    print(f"openpyxl 错误: {e}")

print("\n测试 pandas 使用 openpyxl 引擎读取 .xls 扩展名")
# 模拟一个 .xls 文件（但实际上是 .xlsx 内容）
buffer.seek(0)
try:
    df = pd.read_excel(buffer, engine='openpyxl')
    print("成功")
except Exception as e:
    print(f"失败: {e}")

print("\n检查 xlrd 是否安装")
try:
    import xlrd
    print(f"xlrd 版本: {xlrd.__version__}")
except ImportError:
    print("xlrd 未安装")

# 测试 .xls 文件的引擎自动选择
print("\n测试引擎自动选择（不指定 engine）")
buffer.seek(0)
try:
    df = pd.read_excel(buffer)  # 默认可能会使用 openpyxl，因为内容是 .xlsx
    print("自动选择成功")
except Exception as e:
    print(f"自动选择失败: {e}")